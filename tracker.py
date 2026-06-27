"""Tracks seen job IDs and logs all found jobs to Excel."""
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from scrapers.dou import Job

TRACKER_PATH = Path(__file__).parent / "seen_jobs.xlsx"
SHEET_SEEN = "seen"
SHEET_LOG = "all_jobs"


def _load_workbook():
    if TRACKER_PATH.exists():
        wb = openpyxl.load_workbook(TRACKER_PATH)
        if SHEET_SEEN not in wb.sheetnames:
            _create_seen_sheet(wb)
        if SHEET_LOG not in wb.sheetnames:
            _create_log_sheet(wb)
        return wb

    wb = openpyxl.Workbook()
    wb.active.title = SHEET_SEEN
    _create_seen_sheet(wb)
    _create_log_sheet(wb)
    wb.save(TRACKER_PATH)
    return wb


def _header_style(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E4057")
        cell.alignment = Alignment(horizontal="center")


def _create_seen_sheet(wb):
    if SHEET_SEEN not in wb.sheetnames:
        wb.create_sheet(SHEET_SEEN)
    ws = wb[SHEET_SEEN]
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        _header_style(ws, ["job_id", "title", "company", "url", "source", "found_at"])
    return ws


def _create_log_sheet(wb):
    if SHEET_LOG not in wb.sheetnames:
        wb.create_sheet(SHEET_LOG)
    ws = wb[SHEET_LOG]
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        _header_style(ws, ["Дата", "Назва вакансії", "Компанія", "Джерело", "Посилання", "Опис"])
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 50
        ws.column_dimensions["F"].width = 60
    return ws


def filter_new(jobs):
    seen_ids = set()
    if TRACKER_PATH.exists():
        wb = openpyxl.load_workbook(TRACKER_PATH)
        if SHEET_SEEN in wb.sheetnames:
            ws = wb[SHEET_SEEN]
            seen_ids = {str(row[0].value) for row in ws.iter_rows(min_row=2) if row[0].value}
    result = []
    for j in jobs:
        if j.id not in seen_ids:
            seen_ids.add(j.id)  # deduplicate within current batch too
            result.append(j)
    return result


def mark_seen(job: Job) -> None:
    wb = _load_workbook()

    # Sheet 1: dedup tracker
    ws_seen = wb[SHEET_SEEN]
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M")
    ws_seen.append([job.id, job.title, job.company, job.url, job.source, now])

    # Sheet 2: readable log
    ws_log = wb[SHEET_LOG]
    desc = job.description[:300].replace("\n", " ") if job.description else ""
    ws_log.append([now, job.title, job.company, job.source, job.url, desc])

    wb.save(TRACKER_PATH)
